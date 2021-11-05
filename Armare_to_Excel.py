from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def print_to_excel(proiect, file):
    def write_cell(pos, value, font_value, align=None):
        ws1[str(pos)] = value
        ws1[str(pos)].font = font_value
        ws1[str(pos)].alignment = align

    def merge_cell(start, finish):
        ws1.merge_cells(str(start) +":" + str(finish))

    def add_funct(start, finish):
        return "=SUM(" + str(start) + ":" + str(finish) + ")"

    def product_funct(first, second):
        return "=PRODUCT(" + str(first) + "," + str(second) + ")"


    wb = Workbook()

    dest_filename = file + "\proiect.xlsx"

    ws1 = wb.active
    ws1.title = "Armare Grinzi"

    # fonts
    ftArialCE_11 = Font(name="Arial CE", size=11)
    ftArialCE_10_Bold = Font(name="Arial CE", size=10, bold=True)
    ftSymbol = Font(name="Symbol", size=14)
    ftSymbolBold = Font(name="Symbol", size=10, bold=True)
    ftArial10 = Font(name="Arial", size=10)
    ftArial10Bold = Font(name="Arial", size=10, bold=True)
    AlignmentCenter = Alignment(horizontal="center", vertical="center")

    # end of fonts

    # standard header - position fixed
    write_cell("B2", "Marca", ftArialCE_11, AlignmentCenter)
    write_cell("C2", "Tip Otel", ftArialCE_11, AlignmentCenter)
    write_cell("D2", "f", ftSymbol, AlignmentCenter)
    write_cell("D3", "mm", ftArialCE_11, AlignmentCenter)
    write_cell("E2", "Nr.", ftArialCE_11, AlignmentCenter)
    write_cell("F2", "Nr.", ftArialCE_11, AlignmentCenter)
    write_cell("F3", "buc.", ftArialCE_11, AlignmentCenter)
    write_cell("F4", "tot.", ftArialCE_11, AlignmentCenter)
    write_cell("G2", "Lungime", ftArialCE_11, AlignmentCenter)
    write_cell("G3", "bara", ftArialCE_11, AlignmentCenter)
    write_cell("G4", "m.", ftArialCE_11, AlignmentCenter)
    # end of standard header

    # read and display info from app
    number = 5  # the value of the first blank row

    for item in proiect.dict_grinzi_si_bare.keys():
        write_cell("B"+str(number), "GRINDA"+item.capitalize(), ftArial10Bold)
        number += 1
        for elem in proiect.dict_grinzi_si_bare[item]:
            write_cell("B" + str(number), elem.marca, ftArial10, AlignmentCenter)
            write_cell("C" + str(number), "PC52", ftArial10, AlignmentCenter)
            write_cell("D" + str(number), elem.diam, ftArial10, AlignmentCenter)
            write_cell("E" + str(number),elem.buc, ftArial10, AlignmentCenter)
            write_cell("G" + str(number), elem.lung, ftArial10, AlignmentCenter)
            number += 1
            if proiect.dict_grinzi_si_bare[item].index(elem) == len(proiect.dict_grinzi_si_bare[item]) - 1:
                number += 1
        # end of read from app

    # footer - position relative to info from app
    write_cell("B" + str(number), "Lungime in m/diametru", ftArialCE_11, AlignmentCenter)
    merge_cell("B" + str(number), "F" + str(number))
    number += 1
    write_cell("B" + str(number), "Greutate/m", ftArialCE_11, AlignmentCenter)
    merge_cell("B" + str(number), "F" + str(number))
    number += 1
    write_cell("B" + str(number), "Greutate/diametru", ftArialCE_11, AlignmentCenter)
    merge_cell("B" + str(number), "F" + str(number))
    number += 1
    write_cell("B" + str(number), "TOTAL", ftArialCE_11, AlignmentCenter)
    merge_cell("B" + str(number), "F" + str(number))
    number += 1
    # end of footer

    # get diameters and display
    lista_diam = []

    for elem in proiect.lista_toate_barele:
        lista_diam.append(elem.diam)

    lista_diam = list(set(lista_diam))
    lista_diam.sort()
    number_char = 72

    for elem in lista_diam:
        write_cell(chr(number_char) + str(3), "PC52", ftArialCE_10_Bold, AlignmentCenter)
        write_cell(chr(number_char) + str(4), "f" + str(elem), ftSymbolBold, AlignmentCenter)
        number_char += 1

    merge_cell(chr(72) + str(2), chr(number_char - 1) + str(2))
    # end of diameters display

    # calculation of total lenght per bar
    reinf_weights = {"6": 0.154, "8": 0.395, "10": 0.617, "12": 0.888, "16": 1.578, "20": 2.466}

    for i in range(6, number):
        if type(ws1["D" + str(i)].value) is int:
            for j in range(72, number_char):
                if str(ws1["D" + str(i)].value) in str(ws1[chr(j) + "4"].value):
                    write_cell(chr(j) + str(i), product_funct("G" + str(i), "E" + str(i)),
                               ftArial10, AlignmentCenter)

    # calculation of total weight per diameter
    for i in range(6, number):
        if ws1["B" + str(i)].value == "Lungime in m/diametru":
            for j in range(72, number_char):
                write_cell(chr(j) + str(i), add_funct(chr(j) + str(6), chr(j) + str(number - 5)),
                           ftArial10, AlignmentCenter)

                for k in reinf_weights.keys():
                    if k in str(ws1[chr(int(j)) + str(4)].value):
                        write_cell(chr(j) + str(i + 1), reinf_weights[k], ftArial10, AlignmentCenter)

    # calculation of product between weight and total length per diameter
    for i in range(6, number):
        if ws1["B" + str(i)].value == "Greutate/diametru":
            for j in range(72, number_char):
                write_cell(chr(j) + str(i), product_funct(chr(j) + str(i - 1), chr(j) + str(i - 2)), ftArial10,
                           AlignmentCenter)

    wb.save(filename=dest_filename)


"""
# excel stuff following
def print_to_excel(proiect):
    wb = Workbook()

    dest_filename = "C:\\Users\\bogda\\Desktop\\proiect.xlsx"

    ws1 = wb.active
    ws1.title = "Armare Grinzi"

    # fonts
    ftArialCE_11 = Font(name="Arial CE", size=11)
    ftArialCE_10_Bold = Font(name="Arial CE", size=10, bold=True)
    ftSymbol = Font(name="Symbol", size=14)
    ftSymbolBold = Font(name="Symbol", size=10, bold=True)
    ftArial10 = Font(name="Arial", size=10)
    ftArial10Bold = Font(name="Arial", size=10, bold=True)
    AlignmentCenter = Alignment(horizontal="center", vertical="center")
    # end of fonts

    # standard header - position fixed
    ws1["B2"] = "Marca"
    ws1["C2"] = "Tip Otel"
    ws1["D2"] = "f"
    ws1["D3"] = "mm."
    ws1["E2"] = "Nr."
    ws1["E3"] = "buc."
    ws1["F2"] = "Nr."
    ws1["F3"] = "buc."
    ws1["F4"] = "tot."
    ws1["G2"] = "Lungime"
    ws1["G3"] = "bara"
    ws1["G4"] = "m."

    headers = [ws1["B2"], ws1["C2"], ws1["D3"], ws1["E2"], ws1["E3"], ws1["F2"], ws1["F3"], ws1["F4"], ws1["G2"],
               ws1["G3"], ws1["G4"]]
    for header in headers:
        header.font = ftArialCE_11
        header.alignment = AlignmentCenter

    ws1["D2"].font = ftSymbol
    ws1["D2"].alignment = AlignmentCenter
    # end of standard header

    # read and display info from app
    number = 5  # the value of the first blank row

    for item in proiect.dict_grinzi_si_bare.keys():
        value = "B" + str(number)
        ws1[value] = "GRINDA " + item.capitalize()
        ws1[value].font = ftArial10Bold
        number += 1
        for elem in proiect.dict_grinzi_si_bare[item]:
            value_marca = "B" + str(number)
            value_tip_otel = "C" + str(number)
            value_diam = "D" + str(number)
            value_buc = "E" + str(number)
            value_lung = "G" + str(number)
            ws1[value_marca] = elem.marca
            ws1[value_tip_otel] = "PC52"
            ws1[value_diam] = elem.diam
            ws1[value_buc] = elem.buc
            ws1[value_lung] = elem.lung
            ws1[value_marca].font = ftArial10
            ws1[value_tip_otel].font = ftArial10
            ws1[value_diam].font = ftArial10
            ws1[value_buc].font = ftArial10
            ws1[value_lung].font = ftArial10
            ws1[value_marca].alignment = AlignmentCenter
            ws1[value_tip_otel].alignment = AlignmentCenter
            ws1[value_diam].alignment = AlignmentCenter
            ws1[value_buc].alignment = AlignmentCenter
            ws1[value_lung].alignment = AlignmentCenter
            number += 1
            if proiect.dict_grinzi_si_bare[item].index(elem) == len(proiect.dict_grinzi_si_bare[item]) - 1:
                number += 1
    # end of read from app

    # footer - position relative to info from app
    ws1["B" + str(number)] = "Lungime in m/diametru"
    ws1["B" + str(number)].alignment = AlignmentCenter
    ws1.merge_cells("B" + str(number) + ":F" + str(number))
    ws1["B" + str(number)].font = ftArialCE_11
    number += 1

    ws1["B" + str(number)] = "Greutate/m"
    ws1["B" + str(number)].alignment = AlignmentCenter
    ws1.merge_cells("B" + str(number) + ":F" + str(number))
    ws1["B" + str(number)].font = ftArialCE_11
    number += 1

    ws1["B" + str(number)] = "Greutate/diametru"
    ws1["B" + str(number)].alignment = AlignmentCenter
    ws1.merge_cells("B" + str(number) + ":F" + str(number))
    ws1["B" + str(number)].font = ftArialCE_11
    number += 1

    ws1["B" + str(number)] = "TOTAL"
    ws1["B" + str(number)].alignment = AlignmentCenter
    ws1.merge_cells("B" + str(number) + ":F" + str(number))
    ws1["B" + str(number)].font = ftArialCE_10_Bold
    number += 1
    # end of footer

    # get diameters and display
    lista_diam = []

    for elem in proiect.lista_toate_barele:
        lista_diam.append(elem.diam)

    lista_diam = list(set(lista_diam))
    number_char = 72
    for elem in lista_diam:
        ws1[chr(number_char) + str(3)] = "PC52"
        ws1[chr(number_char) + str(3)].font = ftArialCE_10_Bold
        ws1[chr(number_char) + str(3)].alignment = AlignmentCenter

        ws1[chr(number_char) + str(4)] = "f" + str(elem)
        ws1[chr(number_char) + str(4)].font = ftSymbolBold
        ws1[chr(number_char) + str(4)].alignment = AlignmentCenter
        number_char += 1

    ws1.merge_cells(chr(72) + str(2) + ":" + chr(number_char - 1) + str(2))
    # end of diameters display

    # calculation of total lenght per bar
    reinf_weights = {"6": 0.154, "8": 0.395, "10": 0.617, "12": 0.888, "16": 1.578, "20": 2.466}

    for i in range(6, number):
        if type(ws1["D" + str(i)].value) is int:
            for j in range(72, number_char):
                if str(ws1["D" + str(i)].value) in str(ws1[chr(j) + "4"].value):
                    write_value = chr(j) + str(i)

                    ws1[write_value] = round(ws1["G" + str(i)].value * ws1["E" + str(i)].value,
                                             2)  # change to have formula
                    ws1[write_value].font = ftArial10
                    ws1[write_value].alignment = AlignmentCenter

    # calculation of total weight per diameter
    for i in range(6, number):
        if ws1["B" + str(i)].value == "Lungime in m/diametru":
            for j in range(72, number_char):
                sum_value = chr(j) + str(i)
                weight_value = chr(j) + str(i + 1)
                total_per_diam_value = chr(j) + str(i + 2)

                ws1[sum_value] = "=SUM(" + chr(j) + str(6) + ":" + chr(j) + str(number - 5) + ")"
                ws1[sum_value].font = ftArial10
                ws1[sum_value].alignment = AlignmentCenter
                for k in reinf_weights.keys():
                    if k in str(ws1[chr(int(j)) + str(4)].value):
                        ws1[weight_value] = reinf_weights[k]
                        ws1[weight_value].font = ftArial10
                        ws1[weight_value].alignment = AlignmentCenter

    # calculation of product between weight and total length per diameter
    for i in range(6, number):
        if ws1["B" + str(i)].value == "Greutate/diametru":
            for j in range(72, number_char):
                total_per_diam_value = chr(j) + str(i)
                ws1[total_per_diam_value] = "=PRODUCT(" + chr(j) + str(i - 1) + "," + chr(j) + str(i - 2) + ")"
                ws1[total_per_diam_value].font = ftArial10
                ws1[total_per_diam_value].alignment = AlignmentCenter

    print("completed")
"""

